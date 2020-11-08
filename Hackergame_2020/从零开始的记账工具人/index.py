from openpyxl import load_workbook
from util import to_rmb_upper

import cn2an

wb = load_workbook(filename='bills.xlsx')
ws = wb['Sheet1']

res = 0

for i in range(2, 1002):
    a: str = ws["A%d" % i].value
    b: str = ws["B%d" % i].value

    str = a
    str = str.replace("整", "")
    str = str.replace("佰", "*100+")
    str = str.replace("角", "*0.1+")
    str = str.replace("分", "*0.01+")

    str = str.replace("拾元", "*10+")
    str = str.replace("壹元", "1+")
    str = str.replace("贰元", "2+")
    str = str.replace("叁元", "3+")
    str = str.replace("肆元", "4+")
    str = str.replace("伍元", "5+")
    str = str.replace("陆元", "6+")
    str = str.replace("柒元", "7+")
    str = str.replace("捌元", "8+")
    str = str.replace("玖元", "9+")

    str = str.replace("拾", "*10+")

    str = str.replace("零壹", "1")
    str = str.replace("零贰", "2")
    str = str.replace("零叁", "3")
    str = str.replace("零肆", "4")
    str = str.replace("零伍", "5")
    str = str.replace("零陆", "6")
    str = str.replace("零柒", "7")
    str = str.replace("零捌", "8")
    str = str.replace("零玖", "9")

    str = str.replace("壹", "1")
    str = str.replace("贰", "2")
    str = str.replace("叁", "3")
    str = str.replace("肆", "4")
    str = str.replace("伍", "5")
    str = str.replace("陆", "6")
    str = str.replace("柒", "7")
    str = str.replace("捌", "8")
    str = str.replace("玖", "9")



    str += "0"

    if str[0] == "*":
        str = str[1:]
    # print(a)
    # print(str, eval(str))

    res += eval(str) * int(b)

    print(eval(str))
    # if to_rmb_upper(eval(str)) != a:
    #     if to_rmb_upper(eval(str)) == "壹" + a:
    #         continue
    #     if to_rmb_upper(eval(str)) ==  a + "整":
    #         continue
    #     print(a)
    #     print(str)
    #     print(eval(str))
    #     print(to_rmb_upper(eval(str)), a)

    #     print("=============")


# print(res)
